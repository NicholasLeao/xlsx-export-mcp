#!/usr/bin/env python3
"""XLSX Export MCP Server - Python implementation."""

import json
import sys
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional

from mcp.server.fastmcp import FastMCP
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io

# Export directory configuration
EXPORT_DIR = "/tmp/protex-intelligence-file-exports"


def add_watermark_to_sheet(ws, data_length: int):
    """Add watermark to a worksheet."""
    # Add watermark in first column
    data_end_row = data_length + 1  # +1 for header row
    watermark_row = data_end_row + 2  # Add some space after data
    watermark_col = 1  # First column
    
    # Create watermark cell
    watermark_cell = ws.cell(row=watermark_row, column=watermark_col, value="This content has been generated using Protex Intelligence. The output is intended to assist but may not always be accurate or complete. Please verify important information before acting upon it.")
    
    # Style the watermark
    watermark_cell.font = Font(name='Arial', size=8, color='666666')
    watermark_cell.alignment = Alignment(horizontal='left')


def convert_to_xlsx(
    data: List[Dict[str, Any]], 
    sheet_name: str = "Sheet1",
    headers: Optional[List[str]] = None
) -> bytes:
    """Convert array of objects to XLSX bytes."""
    if not data:
        return b""
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Get headers from first object or use provided headers
    if headers:
        field_names = headers
    else:
        field_names = list(data[0].keys())
    
    # Write headers
    for col_idx, header in enumerate(field_names, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write data rows
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, field_name in enumerate(field_names, 1):
            value = row_data.get(field_name, "")
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Add watermark
    add_watermark_to_sheet(ws, len(data))
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def convert_multi_sheets_to_xlsx(sheets_data: List[Dict[str, Any]]) -> bytes:
    """Convert multiple sheets data to XLSX bytes."""
    if not sheets_data:
        return b""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    for sheet_info in sheets_data:
        sheet_name = sheet_info.get('sheet_name', 'Sheet1')
        data = sheet_info.get('data', [])
        headers = sheet_info.get('headers')
        
        if not data:
            continue
            
        # Create new worksheet
        ws = wb.create_sheet(title=sheet_name)
        
        # Get headers from first object or use provided headers
        if headers:
            field_names = headers
        else:
            field_names = list(data[0].keys())
        
        # Write headers
        for col_idx, header in enumerate(field_names, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Write data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, field_name in enumerate(field_names, 1):
                value = row_data.get(field_name, "")
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add watermark to this sheet
        add_watermark_to_sheet(ws, len(data))
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def get_file_size_string(content: bytes) -> str:
    """Calculate file size string from bytes content."""
    bytes_size = len(content)
    kb = bytes_size / 1024
    
    if kb < 1024:
        return f"{kb:.0f} KB" if kb >= 1 else "1 KB"
    else:
        return f"{kb / 1024:.2f} MB"


async def ensure_export_directory() -> None:
    """Ensure export directory exists, create if it doesn't."""
    export_path = Path(EXPORT_DIR)
    
    if export_path.exists():
        print(f"✓ Export directory exists: {EXPORT_DIR}", file=sys.stderr)
    else:
        try:
            export_path.mkdir(parents=True, exist_ok=True)
            print(f"✓ Created export directory: {EXPORT_DIR}", file=sys.stderr)
        except Exception as e:
            print(f"✗ Failed to create export directory: {e}", file=sys.stderr)
            raise


async def write_xlsx_to_file(xlsx_content: bytes, filename: str) -> str:
    """Write XLSX content to file system."""
    await ensure_export_directory()
    
    filepath = Path(EXPORT_DIR) / filename
    
    try:
        filepath.write_bytes(xlsx_content)
        print(f"✓ File written: {filepath}", file=sys.stderr)
        return str(filepath)
    except Exception as e:
        print(f"✗ Failed to write file: {e}", file=sys.stderr)
        raise


# Create FastMCP server
mcp = FastMCP("xlsx-export-mcp")


@mcp.tool()
async def xlsx_export(
    data: List[Dict[str, Any]],
    filename: str = "output",
    sheet_name: str = "Sheet1",
    description: str = None,
    headers: List[str] = None
) -> dict:
    """Export data to Excel (XLSX) format and save to filesystem.
    
    Args:
        data: Array of objects representing spreadsheet rows
        filename: Filename for the exported file (without extension)
        sheet_name: Name of the worksheet/sheet within the Excel file
        description: Optional description of the file contents
        headers: Optional custom column headers
        
    Returns:
        Dictionary with export results including path and file info
    """
    try:
        # Validate input
        if not data or not isinstance(data, list):
            raise ValueError("Data must be provided as an array of objects")
        
        if len(data) == 0:
            raise ValueError("Data array cannot be empty")
        
        # Convert to XLSX
        xlsx_content = convert_to_xlsx(data, sheet_name, headers)
        
        # Generate UUID and filename
        file_uuid = str(uuid.uuid4())
        sanitized_filename = "".join(c if c.isalnum() or c in "_-" else "_" for c in filename)
        full_filename = f"{sanitized_filename}_{file_uuid}.xlsx"
        file_size = get_file_size_string(xlsx_content)
        row_count = len(data)
        column_count = len(data[0].keys()) if data else 0
        
        # Write XLSX to file system
        filepath = await write_xlsx_to_file(xlsx_content, full_filename)
        
        print(f"✅ XLSX generated: {full_filename} ({file_size})", file=sys.stderr)
        print(f"   Rows: {row_count}, Columns: {column_count}, Sheet: {sheet_name}", file=sys.stderr)
        print(f"   Saved to: {filepath}", file=sys.stderr)
        
        # Return simplified response with essential information
        return {
            "path": full_filename,
            "filetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "filename": full_filename,
            "filesize": file_size,
        }
        
    except Exception as error:
        print(f"Error processing XLSX export: {error}", file=sys.stderr)
        
        return {
            "success": False,
            "error": str(error),
        }


@mcp.tool()
async def xlsx_export_multi_sheet(
    sheets: List[Dict[str, Any]],
    filename: str = "output",
    description: str = None
) -> dict:
    """Export data to multi-sheet Excel (XLSX) format and save to filesystem.
    
    Args:
        sheets: Array of sheet objects, each containing sheet_name, data, and optional headers
        filename: Filename for the exported file (without extension)
        description: Optional description of the file contents
        
    Sheet object format:
        {
            "sheet_name": "Sheet1",
            "data": [{"col1": "val1", "col2": "val2"}, ...],
            "headers": ["col1", "col2"] (optional)
        }
        
    Returns:
        Dictionary with export results including path and file info
    """
    try:
        # Validate input
        if not sheets or not isinstance(sheets, list):
            raise ValueError("Sheets must be provided as an array of sheet objects")
        
        if len(sheets) == 0:
            raise ValueError("At least one sheet must be provided")
        
        # Validate each sheet has required data
        total_rows = 0
        sheet_names = []
        for sheet in sheets:
            if not isinstance(sheet, dict):
                raise ValueError("Each sheet must be an object with sheet_name and data")
            
            if 'data' not in sheet or not isinstance(sheet['data'], list):
                raise ValueError("Each sheet must have a 'data' array")
            
            if len(sheet['data']) == 0:
                raise ValueError("Each sheet's data array cannot be empty")
            
            sheet_name = sheet.get('sheet_name', 'Sheet1')
            sheet_names.append(sheet_name)
            total_rows += len(sheet['data'])
        
        # Convert to multi-sheet XLSX
        print(f"🔄 Generating multi-sheet Excel with {len(sheets)} sheets...", file=sys.stderr)
        xlsx_content = convert_multi_sheets_to_xlsx(sheets)
        
        # Generate UUID and filename
        file_uuid = str(uuid.uuid4())
        sanitized_filename = "".join(c if c.isalnum() or c in "_-" else "_" for c in filename)
        full_filename = f"{sanitized_filename}_{file_uuid}.xlsx"
        file_size = get_file_size_string(xlsx_content)
        
        # Write XLSX to file system
        filepath = await write_xlsx_to_file(xlsx_content, full_filename)
        
        print(f"✅ Multi-sheet XLSX generated: {full_filename} ({file_size})", file=sys.stderr)
        print(f"   Sheets: {len(sheets)} ({', '.join(sheet_names)})", file=sys.stderr)
        print(f"   Total rows: {total_rows}", file=sys.stderr)
        print(f"   Saved to: {filepath}", file=sys.stderr)
        
        # Return simplified response with essential information
        return {
            "path": full_filename,
            "filetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "filename": full_filename,
            "filesize": file_size,
            "sheets": len(sheets),
            "sheet_names": sheet_names
        }
        
    except Exception as error:
        print(f"Error processing multi-sheet XLSX export: {error}", file=sys.stderr)
        
        return {
            "success": False,
            "error": str(error),
        }


def cli_main():
    """CLI entry point."""
    mcp.run()


if __name__ == "__main__":
    cli_main()